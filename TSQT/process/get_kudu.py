from requests import post
import time
from openpyxl import Workbook, styles
from math import fabs
from TSQT.get_cookie import cookie_helper


class GetKudu:
    # start_time_now-->现在开始时间；end_time_now-->现在结束时间；start_time_compare-->对比开始时间； end_time_compare-->对比结束时间；
    def __init__(self, start_time_now, end_time_now, start_time_compare, end_time_compare):
        # 组装大数据cookie
        self.cookie = cookie_helper.composed_cookie('big_query')
        self.X_CSRFToken = self.cookie.split(";")[0].split("=")[1]
        self.big_query_headers = {
            "Host": "bigquery.yunzhangfang.com:8101",
            "X-CSRFToken": self.X_CSRFToken,
            "X-Requested-With": "XMLHttpRequest",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0",
            "Accept": "*/*",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": "http://bigquery.yunzhangfang.com:8101",
            "Referer": "http://bigquery.yunzhangfang.com:8101/hue/editor?editor=86771",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            "Cookie": self.cookie
        }
        self.start_time_now = start_time_now
        self.end_time_now = end_time_now
        self.start_time_compare = start_time_compare
        self.end_time_compare = end_time_compare
        self.big_query_guid = ""
        self.big_query_uuid = ""
        self.big_query_history_id = ""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(("任务类型", "本周", "上个（季报）期", "百分比差值(本周-上期)"))
        # 需要查询的sql
        self.sql_task_num_now = f"select count(*) from fintax_task.task_master where create_time >= '{self.start_time_now}' and create_time <= '{self.end_time_now}'"
        self.sql_task_num_compare = f"select count(*) from fintax_task.task_master where create_time >= '{self.start_time_compare}' and create_time <= '{self.end_time_compare}'"
        self.sql_task_num_dist_now = f"select count(distinct a.qy_id) from fintax_task.task_master a where create_time >= '{self.start_time_now}' and create_time <= '{self.end_time_now}'"
        self.sql_task_num_dist_compare = f"select count(distinct a.qy_id) from fintax_task.task_master a where create_time >= '{self.start_time_compare}' and create_time <= '{self.end_time_compare}'"

    def big_query_kudu(self, sql):
        kudu_data = {
                "notebook": "{\"id\":86685,\"uuid\":\"35480c60-3e1c-4ab1-903c-05e68d728e3a\",\"name\":\"\",\"description\":\"\",\"type\":\"query-kudu\",\"initialType\":\"kudu\",\"coordinatorUuid\":null,\"isHistory\":true,\"isManaged\":false,\"isSaved\":false,\"onSuccessUrl\":null,\"pubSubUrl\":null,\"isPresentationModeDefault\":false,\"isPresentationMode\":false,\"isPresentationModeInitialized\":true,\"presentationSnippets\":{},\"isHidingCode\":false,\"snippets\":[{\"id\":\"9fae589b-fd30-dbd8-2361-8f707c841f39\",\"name\":\"\",\"type\":\"kudu\",\"connector\":{\"optimizer\":\"off\",\"type\":\"kudu\",\"id\":\"kudu\",\"dialect\":\"kudu\",\"is_sql\":true},\"isSqlDialect\":true,\"dialect\":\"kudu\",\"isBatchable\":true,\"autocompleteSettings\":{\"temporaryOnly\":false},\"aceCursorPosition\":{\"row\":0,\"column\":128},\"errors\":[],\"aceErrorsHolder\":[],\"aceWarningsHolder\":[],\"aceErrors\":[],\"aceWarnings\":[],\"editorMode\":true,\"dbSelectionVisible\":false,\"showExecutionAnalysis\":true,\"namespace\":{\"status\":\"CREATED\",\"computes\":[{\"credentials\":{},\"type\":\"direct\",\"id\":\"default\",\"name\":\"default\"}],\"id\":\"default\",\"name\":\"default\"},\"compute\":{\"name\":\"default\",\"namespace\":\"default\",\"id\":\"default\",\"interface\":\"kudu\",\"type\":\"direct\",\"options\":{}},\"database\":\"default\",\"currentQueryTab\":\"queryHistory\",\"pinnedContextTabs\":[],\"loadingQueries\":false,\"queriesHasErrors\":false,\"queriesCurrentPage\":1,\"queriesTotalPages\":1,\"queriesFilter\":\"\",\"queriesFilterVisible\":false,\"statementType\":\"text\",\"statementTypes\":[\"text\",\"file\"],\"statementPath\":\"\",\"externalStatementLoaded\":false,\"associatedDocumentLoading\":true,\"associatedDocumentUuid\":null,\"statement_raw\":\""+sql+"'\",\"statementsList\":[\""+sql+"\"],\"aceSize\":100,\"status\":\"running\",\"statusForButtons\":\"executing\",\"properties\":{},\"viewSettings\":{\"placeHolder\":\"点击侧边栏选择数据源，在这里输入Sql,如：SELECT * FROM 库名.表名\",\"snippetIcon\":\"fa-database\",\"sqlDialect\":true},\"variables\":[],\"hasCurlyBracketParameters\":true,\"variableNames\":[],\"variableValues\":{},\"statement\":\"select count(*) from fintax_task.task_master where create_time >= '"+self.start_time_now+"' and create_time <= '"+self.end_time_now+"'\",\"result\":{\"id\":\"5dfcc941-29a0-79d4-d926-7ec3fa75bd3f\",\"type\":\"table\",\"hasResultset\":true,\"handle\":{},\"meta\":[],\"rows\":null,\"hasMore\":false,\"statement_id\":0,\"statement_range\":{\"start\":{\"row\":0,\"column\":0},\"end\":{\"row\":0,\"column\":0}},\"statements_count\":1,\"previous_statement_hash\":null,\"metaFilter\":{\"query\":\"\",\"facets\":{},\"text\":[]},\"isMetaFilterVisible\":false,\"filteredMetaChecked\":true,\"filteredColumnCount\":-1,\"filteredMeta\":[],\"fetchedOnce\":false,\"startTime\":\"2024-11-15T05:58:48.095Z\",\"endTime\":\"2024-11-15T05:58:48.095Z\",\"executionTime\":0,\"data\":[],\"explanation\":\"\",\"logs\":\"\",\"logLines\":0,\"hasSomeResults\":true},\"showGrid\":true,\"showChart\":false,\"showLogs\":true,\"progress\":0,\"jobs\":[],\"executeNextTimeout\":-1,\"isLoading\":false,\"resultsKlass\":\"results kudu\",\"errorsKlass\":\"results kudu alert alert-error\",\"is_redacted\":false,\"chartType\":\"bars\",\"chartSorting\":\"none\",\"chartScatterGroup\":null,\"chartScatterSize\":null,\"chartScope\":\"world\",\"chartTimelineType\":\"bar\",\"chartLimits\":[5,10,25,50,100],\"chartLimit\":null,\"chartX\":null,\"chartXPivot\":null,\"chartYSingle\":null,\"chartYMulti\":[\"_col0\"],\"chartData\":[],\"chartMapType\":\"marker\",\"chartMapLabel\":null,\"chartMapHeat\":null,\"hideStacked\":true,\"hasDataForChart\":false,\"previousChartOptions\":{\"chartLimit\":null,\"chartX\":null,\"chartXPivot\":null,\"chartYSingle\":null,\"chartMapType\":\"marker\",\"chartMapLabel\":null,\"chartMapHeat\":null,\"chartYMulti\":[\"_col0\"],\"chartScope\":\"world\",\"chartTimelineType\":\"bar\",\"chartSorting\":\"none\",\"chartScatterGroup\":null,\"chartScatterSize\":null},\"isResultSettingsVisible\":false,\"settingsVisible\":false,\"checkStatusTimeout\":null,\"getLogsTimeout\":null,\"topRisk\":null,\"suggestion\":\"\",\"hasSuggestion\":null,\"compatibilityCheckRunning\":false,\"compatibilitySourcePlatforms\":[{\"name\":\"Teradata\",\"value\":\"teradata\"},{\"name\":\"Oracle\",\"value\":\"oracle\"},{\"name\":\"Netezza\",\"value\":\"netezza\"},{\"name\":\"Impala\",\"value\":\"impala\"},{\"name\":\"Hive\",\"value\":\"hive\"},{\"name\":\"DB2\",\"value\":\"db2\"},{\"name\":\"Greenplum\",\"value\":\"greenplum\"},{\"name\":\"MySQL\",\"value\":\"mysql\"},{\"name\":\"PostgreSQL\",\"value\":\"postgresql\"},{\"name\":\"Informix\",\"value\":\"informix\"},{\"name\":\"SQL Server\",\"value\":\"sqlserver\"},{\"name\":\"Sybase\",\"value\":\"sybase\"},{\"name\":\"Access\",\"value\":\"access\"},{\"name\":\"Firebird\",\"value\":\"firebird\"},{\"name\":\"ANSISQL\",\"value\":\"ansisql\"},{\"name\":\"Generic\",\"value\":\"generic\"}],\"compatibilityTargetPlatforms\":[{\"name\":\"Impala\",\"value\":\"impala\"},{\"name\":\"Hive\",\"value\":\"hive\"}],\"showOptimizer\":false,\"wasBatchExecuted\":false,\"isReady\":true,\"lastExecuted\":1731650328090,\"lastAceSelectionRowOffset\":0,\"executingBlockingOperation\":null,\"showLongOperationWarning\":false,\"lastExecutedStatements\":\""+sql+"\",\"lastExecutedSelectionRange\":{\"start\":{\"row\":0,\"column\":128},\"end\":{\"row\":0,\"column\":128}},\"formatEnabled\":true,\"isFetchingData\":false,\"isCanceling\":false,\"aceAutoExpand\":false,\"lastCheckStatusRequest\":{\"readyState\":4,\"responseText\":\"{\\\"status\\\": 0, \\\"query_status\\\": {\\\"status\\\": \\\"available\\\"}}\",\"responseJSON\":{\"status\":0,\"query_status\":{\"status\":\"available\"}},\"status\":200,\"statusText\":\"OK\"},\"lastGetLogsRequest\":{\"readyState\":4,\"responseText\":\"{\\\"status\\\": 0, \\\"progress\\\": 100, \\\"jobs\\\": [], \\\"logs\\\": \\\"\\\", \\\"isFullLogs\\\": true}\",\"responseJSON\":{\"status\":0,\"progress\":100,\"jobs\":[],\"logs\":\"\",\"isFullLogs\":true},\"status\":200,\"statusText\":\"OK\"}}],\"selectedSnippet\":\"kudu\",\"creatingSessionLocks\":[],\"sessions\":[{\"type\":\"kudu\",\"properties\":[],\"id\":null}],\"directoryUuid\":\"\",\"dependentsCoordinator\":[],\"historyFilter\":\"\",\"historyFilterVisible\":false,\"loadingHistory\":false,\"historyInitialHeight\":4664,\"forceHistoryInitialHeight\":true,\"historyCurrentPage\":1,\"historyTotalPages\":64,\"schedulerViewModel\":null,\"schedulerViewModelIsLoaded\":false,\"isBatchable\":true,\"isExecutingAll\":false,\"executingAllIndex\":0,\"retryModalConfirm\":null,\"retryModalCancel\":null,\"unloaded\":false,\"updateHistoryFailed\":false,\"viewSchedulerId\":\"\",\"loadingScheduler\":false}",
                "snippet": "{\"id\":\"9fae589b-fd30-dbd8-2361-8f707c841f39\",\"type\":\"kudu\",\"status\":\"running\",\"statementType\":\"text\",\"statement\":\""+sql+"\",\"aceCursorPosition\":{\"row\":0,\"column\":128},\"statementPath\":\"\",\"associatedDocumentUuid\":null,\"properties\":{},\"result\":{\"id\":\"5dfcc941-29a0-79d4-d926-7ec3fa75bd3f\",\"type\":\"table\",\"handle\":{}},\"database\":\"default\",\"compute\":{\"name\":\"default\",\"namespace\":\"default\",\"id\":\"default\",\"interface\":\"kudu\",\"type\":\"direct\",\"options\":{}},\"wasBatchExecuted\":false}"
        }
        url = 'http://bigquery.yunzhangfang.com:8101/notebook/api/execute/kudu'
        resp = post(url=url, headers=self.big_query_headers, data=kudu_data)
        print(resp.text)
        try:
            self.big_query_guid = resp.json()['handle']['guid']
            self.big_query_uuid = resp.json()['history_uuid']
            self.big_query_history_id = resp.json()['history_id']
        except KeyError as e:
            print("大数据查询出错，请检查cookie" + str(e))
        return self.big_query_fetch_data(sql)


    def big_query_fetch_data(self, sql):
        headers = {
                "Host": "bigquery.yunzhangfang.com:8101",
                "X-CSRFToken": self.X_CSRFToken,
                "X-Requested-With": "XMLHttpRequest",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0",
                "Accept": "text/plain, */*; q=0.01",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "Origin": "http://bigquery.yunzhangfang.com:8101",
                "Referer": "http://bigquery.yunzhangfang.com:8101/hue/editor?editor="+str(self.big_query_history_id),
                "Accept-Encoding": "gzip, deflate",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
                "Cookie": self.cookie
        }
        fetch_data = {
              "notebook": "{\"id\":"+str(self.big_query_history_id)+",\"uuid\":\""+str(self.big_query_uuid)+"\",\"isSaved\":false,\"sessions\":[{\"type\":\"kudu\",\"properties\":[],\"id\":null}],\"type\":\"query-kudu\",\"name\":\"\"}",
              "snippet": "{\"id\":\"ff0af6fd-fcda-6137-97d9-4b31f2d5936a\",\"type\":\"kudu\",\"status\":\"available\",\"statementType\":\"text\",\"statement\":\""+sql+"\",\"aceCursorPosition\":{\"column\":145,\"row\":0},\"statementPath\":\"\",\"associatedDocumentUuid\":null,\"properties\":{},\"result\":{\"id\":\""+str(self.big_query_uuid)+"\",\"type\":\"table\",\"handle\":{\"sync\":false,\"has_result_set\":true,\"result\":{\"has_more\":true,\"type\":\"table\",\"meta\":[{\"comment\":\"\",\"type\":\"STRING_TYPE\",\"name\":\"_col0\"}],\"data\":[]},\"statement\":\"-- {\\\"username\\\":\\\"wangxue1@yunzhangfang.com\\\",\\\"appid\\\":\\\"hue-yzf-desktop\\\"}\\n"+sql+"\",\"modified_row_count\":0,\"guid\":\""+str(self.big_query_guid)+"\"}},\"database\":\"default\",\"compute\":{\"name\":\"default\",\"namespace\":\"default\",\"id\":\"default\",\"interface\":\"kudu\",\"type\":\"direct\",\"options\":{}},\"wasBatchExecuted\":false}",
              "rows": "100",
              "startOver": "true"
        }
        url = 'http://bigquery.yunzhangfang.com:8101/notebook/api/fetch_result_data'
        resp = post(url=url, headers=headers, data=fetch_data)
        try:
            result = resp.json()['result']['data'][0][0]
            return result
        except Exception as e:
            print("json转换有误" + str(e))

    # 数据分析，对比本期和上期数据
    def analysis_to_excel(self, num_now, num_compare, num_dist_now, num_dist_compare):
        temp_total_task = round((fabs(num_now - num_compare) / num_compare)*100, 2)
        temp_active_company = round((fabs(num_dist_now - num_dist_compare) / num_dist_compare)*100, 2)
        self.ws.append(("总任务量", num_now, num_compare, temp_total_task))
        self.ws.append(("活跃企业数", num_dist_now, num_dist_compare, temp_active_company))
        # 本期比上期总任务量多3%
        if num_now > num_compare and temp_total_task > 3:
            self.ws.cell(2, 4).fill = styles.PatternFill(start_color='68eca3', end_color='68eca3', fill_type='solid')
            self.ws.cell(2, 4).value = f"{temp_total_task}% ↑"
        # 本期比上期总任务量少3%
        elif num_now < num_compare and temp_total_task > 3:
            self.ws.cell(2, 4).fill = styles.PatternFill(start_color='ef7171', end_color='ef7171', fill_type='solid')
            self.ws.cell(2, 4).value = f"{temp_total_task}% ↓"
        # 本期比上期总活跃企业数多3%
        if num_dist_now > num_dist_compare and temp_total_task > 3:
            self.ws.cell(3, 4).fill = styles.PatternFill(start_color='68eca3', end_color='68eca3', fill_type='solid')
            self.ws.cell(3, 4).value = f"{temp_active_company}% ↑"
        # 本期比上期总活跃企业数少3%
        elif num_dist_now < num_dist_compare and temp_total_task > 3:
            self.ws.cell(3, 4).fill = styles.PatternFill(start_color='ef7171', end_color='ef7171', fill_type='solid')
            self.ws.cell(3, 4).value = f"{temp_active_company}% ↓"

    # 保存xlsx文件
    def save_file(self):
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[:10]
        self.wb.save(f'./kudu_{current_time}.xlsx')

    # 任务主函数
    def task_process(self):
        num_now = self.big_query_kudu(self.sql_task_num_now)
        num_compare = self.big_query_kudu(self.sql_task_num_compare)
        num_dist_now = self.big_query_kudu(self.sql_task_num_dist_now)
        num_dist_compare = self.big_query_kudu(self.sql_task_num_dist_compare)
        return num_now, num_compare, num_dist_now, num_dist_compare


if __name__ == '__main__':
    kudu = GetKudu('2024-12-05 00:00:00', '2024-12-11 23:59:59', '2024-11-05 00:00:00', '2024-11-11 23:59:59')
    num_now, num_compare, num_dist_now, num_dist_compare = kudu.task_process()
    kudu.analysis_to_excel(num_now, num_compare, num_dist_now, num_dist_compare)
    kudu.save_file()


