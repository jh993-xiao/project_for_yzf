import requests
from enum import Enum
from openpyxl import Workbook, styles
import math
import time


# 类型枚举类  jbxx-->基本信息；swcsh-->税务初始化；bsjc-->报税检查；sbrw-->申报任务；fprzcj-->发票认证采集
class TypeEnum(Enum):
    nothing = None
    fprzcj = 20
    sbrw = 5
    bsjc = 9
    swcsh = 10
    jbxx = 1


class GetData:
    # start_time_now-->现在开始时间；end_time_now-->现在结束时间；start_time_compare-->对比开始时间； end_time_compare-->对比结束时间；
    def __init__(self, start_time_now, end_time_now, start_time_compare, end_time_compare, cookie, task_type=None):
        # 数据初始化
        self.task_type_dict = {
            "None": "全部",
            "1": "基本信息",
            "5": "申报任务",
            "9": "报税检查",
            "10": "税务初始化",
            "20": "发票认证采集"
        }
        self.start_time_now = start_time_now
        self.end_time_now = end_time_now
        self.start_time_compare = start_time_compare
        self.end_time_compare = end_time_compare
        self.data = {
            "beginDate": start_time_now,
            "endDate": end_time_now,
            "rwlx": "",
            "szType": "sj"
        }
        self.last_quarter_data = {
            "beginDate": start_time_compare,
            "endDate": end_time_compare,
            "rwlx": "",
            "szType": "sj"
        }
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(("任务类型", "任务结果", "本周", "上个（季报）期", "差值(本周-上期)"))
        self.headers = {
            "Cookie": cookie
        }
        self.task_type = task_type
        self.task_process_list = [types.value for types in TypeEnum]

    def task_status(self, data):
        url = "https://taxsupport.yunzhangfang.com/tax/dash/codeCategoryPie"
        resp = requests.get(url=url, params=data, headers=self.headers)
        return resp

    # 拿到的值为各个类型的值，需要计算所占百分比
    def data_analysis(self, data_list):
        total = 0
        for index, data in enumerate(data_list):
            total += data['task_qty']
        for index, data in enumerate(data_list):
            data['task_qty'] = round(data['task_qty']/total, 4)
        return data_list

    # 与上一个季度或者报税期做数据对比
    def compare_last_quarter(self):
        temp = []
        # 标记所处第几行单元格
        count = 1
        for item in self.ws.iter_cols(min_col=3, max_col=4, values_only=True):
            temp.append(item)
        for value in temp[0][1:]:
            self.ws.cell(count + 1, 3).value = f"{value}%"
            # 现在比对比百分比多3%以上，对文件进行处理（包括数据回写，颜色改变）
            if value > temp[1][count] and math.fabs(value - temp[1][count]) > 3:
                self.ws.cell(count + 1, 3).value = f"{value}% ↑"
                if '成功' in self.ws.cell(count+1, 2).value:
                    self.ws.cell(count+1, 3).fill = styles.PatternFill(start_color='68eca3',
                                                                                end_color='68eca3', fill_type='solid')
                else:
                    self.ws.cell(count + 1, 3).fill = styles.PatternFill(start_color='ef7171',
                                                                                  end_color='ef7171', fill_type='solid')
            # 现在比对比百分比少3%以上，对文件进行处理（包括数据回写，颜色改变）
            if value < temp[1][count] and math.fabs(value - temp[1][count]) > 3:
                self.ws.cell(count + 1, 3).value = f"{value}% ↓"
                if '成功' in self.ws.cell(count+1, 2).value:
                    self.ws.cell(count + 1, 3).fill = styles.PatternFill(start_color='ef7171',
                                                                                  end_color='ef7171', fill_type='solid')

                else:
                    self.ws.cell(count + 1, 3).fill = styles.PatternFill(start_color='68eca3',
                                                                                  end_color='68eca3', fill_type='solid')
            self.ws.cell(count + 1, 5).value = f"{round(value - temp[1][count], 2)}%"
            self.ws.cell(count + 1, 4).value = f"{temp[1][count]}%"
            count += 1

    # 保存对比后的文件
    def save_file(self):
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[:10]
        self.wb.save(f'./result_{current_time}.xlsx')

    # 主任务进度
    def task_process(self):
        for index, task_type in enumerate(self.task_process_list):
            self.write_to_excel(task_type)
        # self.compare_last_quarter() #执行放入界面

    # 将现在和上期（季度）数据写入文件，此时还没做数据对比
    def write_to_excel(self, task_type):
        # 分子，分母
        temp_up = ()
        temp_down = ()

        self.data['rwlx'] = task_type
        resp = self.task_status(self.data)
        data_list = self.data_analysis(resp.json()['data'])

        self.last_quarter_data['rwlx'] = task_type
        last_quarter_resp = self.task_status(self.last_quarter_data)
        last_quarter_data_list = self.data_analysis(last_quarter_resp.json()['data'])

        for index, last_data in enumerate(last_quarter_data_list):
            for i, data in enumerate(data_list):
                if last_data["categoryName"] == data["categoryName"]:
                    data['last_task_qty'] = last_data['task_qty']
        # 过滤楚不包含用户问题的成功率
        for index, data in enumerate(data_list):
            if data['categoryName'] == '成功':
                temp_up = (data['task_qty'], data['last_task_qty'])
            if data['categoryName'] == '用户问题':
                temp_down = (data['task_qty'], data['last_task_qty'])

        data_list.append({'codeCagegory': '100', 'categoryName': '成功（除用户问题）', 'task_qty': round(temp_up[0]/(1-temp_down[0]), 4), 'last_task_qty': round(temp_up[1]/(1-temp_down[1]), 4)})
        sort_list = ['成功', '成功（除用户问题）', '系统问题', '税局问题', '用户问题']
        data_list_sort = []
        for index, item in enumerate(sort_list):
            for i, data in enumerate(data_list):
                if item == data['categoryName']:
                    data_list_sort.append(data)

        # 过滤多余不用写入的数据
        for index, data in enumerate(data_list_sort):
            if data['codeCagegory'] in ['0', '1', '2', '3', '100']:
                self.ws.append((self.task_type_dict.get(str(task_type)), data['categoryName'], round(data['task_qty']*100, 2), round(data['last_task_qty']*100, 2)))


if __name__ == '__main__':
    get_data = GetData("2024-12-01 00:00:00", "2024-12-04 23:59:59", "2024-11-01 00:00:00", "2024-11-04 23:59:59", "33ea3b9c8d374d20a81450c44a0c2f81=WyIzMjEwMDA1OTU5Il0; dd_access_token=eyJhbGciOiJIUzI1NiJ9.eyJwaG9uZSI6IiIsImdzSWQiOiJkZWZhdWx0IiwiaXNzIjoiYXV0aDAiLCJ1c2VyTmFtZSI6ImNJSG9VNUpOSUNNVzBYamlpc1ZvUU5naUVpRSIsImV4cCI6MTczMTg5NjQzMiwidXNlcklkIjoiMjMwNTc5MTUzIn0.fSWyA0uF0WKTapACl2zsedHLSNn5YZAQYcSOuCPZkLs; dd_refresh_token=a80986dce48a4c47aae4980f4de8bcbb; userInfo={%22userId%22:1942%2C%22username%22:%22%E7%BA%AA%E8%B1%AA%22%2C%22email%22:%22jihao@yunzhangfang.com%22}")
    get_data.task_process()
    get_data.compare_last_quarter()
    get_data.save_file()

